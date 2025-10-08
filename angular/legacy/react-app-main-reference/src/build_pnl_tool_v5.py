#!/usr/bin/env python3
"""
Builds the Liberty Tax P&L Tool v0.5 (macro-free) using openpyxl.

Output: dist/LT_PnL_Tool_v0.5.xlsx
"""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference, DoughnutChart
import os

# Brand Colors (following Liberty Tax brand guidelines)
LIBERTY_BLUE = "002D72"    # Liberty Navy (same as web app)
LIBERTY_RED = "EA0029"     # Liberty Red (primary brand color)
ACCENT_GREY = "F2F2F2"
BORDER_GREY = "DDDDDD"

# Brand Typography for Excel compatibility
BRAND_FONT = "Arial"       # Primary font (Proxima Nova fallback for Excel)

GREEN = "C6EFCE"   # fill for green
YELLOW = "FFEB9C"
RED = "F4CCCC"

def thin_border():
    return Border(left=Side(style="thin", color=BORDER_GREY),
                  right=Side(style="thin", color=BORDER_GREY),
                  top=Side(style="thin", color=BORDER_GREY),
                  bottom=Side(style="thin", color=BORDER_GREY))

def brand_font(bold=False, size=10):
    """Create brand-consistent font using Arial (Proxima Nova fallback for Excel)"""
    return Font(name=BRAND_FONT, bold=bold, size=size)

def brand_header_font(size=12):
    """Brand header font with semibold equivalent styling"""
    return Font(name=BRAND_FONT, bold=True, size=size, color=LIBERTY_BLUE)

def header(cell):
    cell.font = brand_font(bold=True)  # Use brand-consistent font
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

def build_wb():
    wb = Workbook()
    wb.remove(wb.active)

    # -------------------- Welcome --------------------
    ws_w = wb.create_sheet("Welcome")
    ws_w["A1"] = "Welcome – Quick Start"
    ws_w["A1"].font = brand_header_font(size=14)  # Brand header typography

    ws_w["A3"] = "Region"
    ws_w["B3"] = "U.S."
    dv_region = DataValidation(type="list", formula1='"U.S.,Canada"')
    ws_w.add_data_validation(dv_region); dv_region.add(ws_w["B3"])

    prompts = [
        ("Planned return count (2026)", 1600),
        ("Average net fee (ANF) $", 125),
        ("Discounts % of Gross Fees", 3),
        ("Target Net Margin % (green)", 20),
        ("Cost/Return green threshold $", 25),
        ("Cost/Return yellow upper bound $", 35),
    ]
    r = 5
    for label, default in prompts:
        ws_w[f"A{r}"] = label; ws_w[f"B{r}"] = default
        header(ws_w[f"A{r}"])
        ws_w[f"A{r}"].alignment = Alignment(vertical="center")
        r += 2

    # Navigation
    ws_w["D2"] = "Go to Inputs"; ws_w["D2"].hyperlink = "#'Inputs'!A1"; ws_w["D2"].style = "Hyperlink"
    ws_w["E2"] = "Go to Dashboard"; ws_w["E2"].hyperlink = "#'Dashboard'!A1"; ws_w["E2"].style = "Hyperlink"

    for col in range(1, 6):
        ws_w.column_dimensions[get_column_letter(col)].width = 42
    ws_w.sheet_view.showGridLines = False

    # -------------------- Inputs --------------------
    ws_in = wb.create_sheet("Inputs")
    ws_in["A1"] = "Scenario"; ws_in["B1"] = "Custom"
    dv = DataValidation(type="list", formula1='"Custom,Good,Better,Best"')
    ws_in.add_data_validation(dv); dv.add(ws_in["B1"])
    ws_in["D1"] = "Region"; ws_in["E1"] = "=Welcome!B3"

    ws_in["A2"] = "Income Drivers"; header(ws_in["A2"])
    drivers = [
        ("Average Net Fee ($)", 125, 130, 135, 140),
        ("Tax Prep Returns (#)", 1600, 1680, 1840, 2000),
        ("TaxRush Returns (#)", 0, 0, 0, 0),
    ]
    row = 3
    for i, (name, default, good, better, best) in enumerate(drivers, start=2):
        ws_in[f"A{row}"] = name
        if name == "Average Net Fee ($)":
            scen = f'=IF($B$1="Custom",{default},INDEX(Presets!$B${i}:$D${i}, MATCH($B$1,Presets!$B$1:$D$1,0)))'
            ws_in[f"B{row}"] = f'=IF(ISBLANK(Welcome!B7),{scen},Welcome!B7)'
        elif name == "Tax Prep Returns (#)":
            scen = f'=IF($B$1="Custom",{default},INDEX(Presets!$B${i}:$D${i}, MATCH($B$1,Presets!$B$1:$D$1,0)))'
            ws_in[f"B{row}"] = f'=IF(ISBLANK(Welcome!B5),{scen},Welcome!B5)'
        else:  # TaxRush
            scen = f'=IF($B$1="Custom",{default},INDEX(Presets!$B${i}:$D${i}, MATCH($B$1,Presets!$B$1:$D$1,0)))'
            ws_in[f"B{row}"] = f'=IF(Inputs!$E$1="U.S.",0,{scen})'
        row+=1

    row+=1
    ws_in[f"A{row}"] = "Expense Percentages"; header(ws_in[f"A{row}"]); row+=1
    expenses = [
        ("Customer Discounts (% of Gross Fees)", 3, 3, 3, 3),
        ("Salaries (% of Gross Fees)", 25, 26, 24, 22),
        ("Rent (% of Gross Fees)", 18, 18, 17, 16),
        ("Office Supplies (% of Gross Fees)", 3.5, 3.5, 3.5, 3.5),
        ("Tax Prep Royalties (% of Tax Prep Income)", 14, 14, 14, 14),
        ("Advertising Royalties (% of Tax Prep Income)", 5, 5, 5, 5),
        ("Other Misc/Shortages (% of TP Income)", 2.5, 2.5, 2.5, 2.5),
    ]
    start_idx = 5
    for j, (name, default, good, better, best) in enumerate(expenses, start=start_idx):
        ws_in[f"A{row}"] = name
        if name == "Customer Discounts (% of Gross Fees)":
            scen = f'=IF($B$1="Custom",{default},INDEX(Presets!$B${j}:$D${j}, MATCH($B$1,Presets!$B$1:$D$1,0)))'
            ws_in[f"B{row}"] = f'=IF(ISBLANK(Welcome!B9),{scen},Welcome!B9)'
        else:
            ws_in[f"B{row}"] = f'=IF($B$1="Custom",{default},INDEX(Presets!$B${j}:$D${j}, MATCH($B$1,Presets!$B$1:$D$1,0)))'
        row+=1

    row+=1
    ws_in[f"A{row}"] = "Thresholds"; header(ws_in[f"A{row}"]); row+=1
    ws_in[f"A{row}"] = "Cost per Return – Green if ≤"; ws_in[f"B{row}"] = "=Welcome!B11"; row+=1
    ws_in[f"A{row}"] = "Cost per Return – Yellow upper bound"; ws_in[f"B{row}"] = "=Welcome!B13"; row+=1
    ws_in[f"A{row}"] = "Net Income Margin – Green if ≥ (%)"; ws_in[f"B{row}"] = "=Welcome!B9"; row+=1
    ws_in[f"A{row}"] = "Net Income Margin – Yellow lower bound (%)"; ws_in[f"B{row}"] = 10; row+=1
    ws_in[f"A{row}"] = "Net Income – Red if ≤ (absolute $)"; ws_in[f"B{row}"] = -5000; row+=1

    for col in range(1, 6): ws_in.column_dimensions[get_column_letter(col)].width = 44
    ws_in.sheet_view.showGridLines = False
    ws_in.freeze_panes = "A3"

    # -------------------- Presets --------------------
    ws_p = wb.create_sheet("Presets")
    ws_p["A1"] = "Parameter"; ws_p["B1"] = "Good"; ws_p["C1"] = "Better"; ws_p["D1"] = "Best"
    ws_p["A1"].font = brand_font(bold=True)
    for c in ["B","C","D"]: ws_p[f"{c}1"].font = brand_font(bold=True)
    r = 2
    for tup in drivers + expenses:
        n, d, g, btr, bst = tup
        ws_p[f"A{r}"]=n; ws_p[f"B{r}"]=g; ws_p[f"C{r}"]=btr; ws_p[f"D{r}"]=bst; r+=1
    for c in range(1,5): ws_p.column_dimensions[get_column_letter(c)].width=44
    ws_p.sheet_view.showGridLines = False

    # -------------------- Results --------------------
    ws_r = wb.create_sheet("Results")
    ws_r["A1"], ws_r["B1"] = "Metric", "Value"
    ws_r["A1"].font = ws_r["B1"].font = brand_font(bold=True)

    def ref(label):
        for rr in range(1, ws_in.max_row+1):
            if ws_in[f"A{rr}"].value == label:
                return f"Inputs!B{rr}"
        return ""

    avg=ref("Average Net Fee ($)"); ret=ref("Tax Prep Returns (#)"); tpr=ref("TaxRush Returns (#)")
    discp=ref("Customer Discounts (% of Gross Fees)")
    salp=ref("Salaries (% of Gross Fees)"); rentp=ref("Rent (% of Gross Fees)"); supp=ref("Office Supplies (% of Gross Fees)")
    royp=ref("Tax Prep Royalties (% of Tax Prep Income)"); advp=ref("Advertising Royalties (% of Tax Prep Income)"); miscp=ref("Other Misc/Shortages (% of TP Income)")

    r=2
    ws_r[f"A{r}"]="Gross Fees"; ws_r[f"B{r}"]=f"={avg}*{ret}"; gross=f"Results!B{r}"; r+=1
    ws_r[f"A{r}"]="Customer Discounts"; ws_r[f"B{r}"]=f"={gross}*{discp}/100"; disc=f"Results!B{r}"; r+=1
    ws_r[f"A{r}"]="Tax Prep Income"; ws_r[f"B{r}"]=f"={gross}-{disc}"; income=f"Results!B{r}"; r+=1
    ws_r[f"A{r}"]="Total Returns"; ws_r[f"B{r}"]=f"={ret}+{tpr}"; total=f"Results!B{r}"; r+=1
    ws_r[f"A{r}"]="Salaries"; ws_r[f"B{r}"]=f"={gross}*{salp}/100"; sal=f"Results!B{r}"; r+=1
    ws_r[f"A{r}"]="Rent"; ws_r[f"B{r}"]=f"={gross}*{rentp}/100"; rent=f"Results!B{r}"; r+=1
    ws_r[f"A{r}"]="Office Supplies"; ws_r[f"B{r}"]=f"={gross}*{supp}/100"; sup=f"Results!B{r}"; r+=1
    ws_r[f"A{r}"]="Royalties"; ws_r[f"B{r}"]=f"={income}*{royp}/100"; roy=f"Results!B{r}"; r+=1
    ws_r[f"A{r}"]="Advertising Royalties"; ws_r[f"B{r}"]=f"={income}*{advp}/100"; adv=f"Results!B{r}"; r+=1
    ws_r[f"A{r}"]="Misc / Shortages"; ws_r[f"B{r}"]=f"={income}*{miscp}/100"; misc=f"Results!B{r}"; r+=1
    ws_r[f"A{r}"]="Total Expenses"; ws_r[f"B{r}"]=f"={sal}+{rent}+{sup}+{roy}+{adv}+{misc}"; texp=f"Results!B{r}"; r+=1
    ws_r[f"A{r}"]="Net Income"; ws_r[f"B{r}"]=f"={income}-{texp}"; net=f"Results!B{r}"; r+=1
    ws_r[f"A{r}"]="Cost per Return"; ws_r[f"B{r}"]=f"={texp}/{total}"; cpr=f"Results!B{r}"; r+=1
    ws_r[f"A{r}"]="Net Income Margin (%)"; ws_r[f"B{r}"]=f"={net}/{income}*100"; nim=f"Results!B{r}"

    for c in range(1,3): ws_r.column_dimensions[get_column_letter(c)].width=42
    ws_r.sheet_view.showGridLines = False
    ws_r.freeze_panes = "A2"

    # Threshold refs
    cpr_g=ref("Cost per Return – Green if ≤"); cpr_y=ref("Cost per Return – Yellow upper bound")
    nim_g=ref("Net Income Margin – Green if ≥ (%)"); nim_y=ref("Net Income Margin – Yellow lower bound (%)")
    ni_r =ref("Net Income – Red if ≤ (absolute $)")

    # Conditional Formats
    green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    red_fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")

    ws_r.conditional_formatting.add(cpr, CellIsRule(operator="lessThanOrEqual", formula=[cpr_g], fill=green_fill))
    ws_r.conditional_formatting.add(cpr, CellIsRule(operator="greaterThan", formula=[cpr_y], fill=red_fill))
    ws_r.conditional_formatting.add(cpr, FormulaRule(formula=[f"AND({cpr}>{cpr_g},{cpr}<={cpr_y})"], fill=yellow_fill))

    ws_r.conditional_formatting.add(nim, CellIsRule(operator="greaterThanOrEqual", formula=[nim_g], fill=green_fill))
    ws_r.conditional_formatting.add(nim, CellIsRule(operator="lessThan", formula=[nim_y], fill=red_fill))
    ws_r.conditional_formatting.add(nim, FormulaRule(formula=[f"AND({nim}>={nim_y},{nim}<{nim_g})"], fill=yellow_fill))

    ws_r.conditional_formatting.add(net, CellIsRule(operator="lessThanOrEqual", formula=[ni_r], fill=red_fill))

    # -------------------- Dashboard --------------------
    ws_d = wb.create_sheet("Dashboard")
    ws_d["A1"] = "Liberty Tax – KPI Dashboard"; ws_d["A1"].font = brand_header_font(size=14)
    ws_d.merge_cells("A1:E1")

    # KPI rows
    ws_d["A3"] = "Net Income";   ws_d["B3"] = f"={net}"
    ws_d["A6"] = "Net Margin %"; ws_d["B6"] = f"={nim}"
    ws_d["A9"] = "Cost per Return"; ws_d["B9"] = f"={cpr}"

    for r in [3,6,9]:
        ws_d[f"A{r}"].font = brand_font(bold=True)
        ws_d[f"A{r}"].fill = PatternFill(start_color=ACCENT_GREY, end_color=ACCENT_GREY, fill_type="solid")
        ws_d[f"A{r}"].border = thin_border()
        ws_d[f"B{r}"].border = thin_border()

    # Indicators & status
    ws_d["C2"] = "Indicator"; ws_d["D2"] = "Status"
    ws_d["C2"].font = ws_d["D2"].font = brand_font(bold=True)

    ws_d["C3"] = '=IF(B3<=Inputs!$B$23,"🔴","🟢")'
    ws_d["D3"] = '=IF(B3<=Inputs!$B$23,"Below target","OK/Above")'

    ws_d["C6"] = f'=IF({nim}>={nim_g},"🟢",IF({nim}>={nim_y},"🟡","🔴"))'
    ws_d["D6"] = f'=IF({nim}>={nim_g},"Best",IF({nim}>={nim_y},"Good","Needs Attention"))'

    ws_d["C9"] = f'=IF({cpr}<={cpr_g},"🟢",IF({cpr}<={cpr_y},"🟡","🔴"))'
    ws_d["D9"] = f'=IF({cpr}<={cpr_g},"Best",IF({cpr}<={cpr_y},"Good","Needs Attention"))'

    # Color symbol cells
    def color_symbol(cell):
        ws_d.conditional_formatting.add(cell, FormulaRule(formula=[f'LEFT({cell},1)="🔴"'], fill=red_fill))
        ws_d.conditional_formatting.add(cell, FormulaRule(formula=[f'LEFT({cell},1)="🟡"'], fill=yellow_fill))
        ws_d.conditional_formatting.add(cell, FormulaRule(formula=[f'LEFT({cell},1)="🟢"'], fill=green_fill))

    for rr in [3,6,9]: color_symbol(f"C{rr}")

    # Mini Practice Progress (5 segments) in E3:I3
    ws_d["E2"]="Practice Progress"; ws_d["E2"].font = brand_font(bold=True)
    ws_d["J2"]="Completed"; ws_d["J3"]='=Practice!B20'  # mirror count
    # paint segment cells with rules
    for i in range(5):
        col = 5 + i  # E..I
        cell = f"{get_column_letter(col)}3"
        ws_d[cell] = ""  # visual segment
        ws_d[cell].border = thin_border()
        # Conditional fills based on completion count (Practice!B20)
        # green if count >= i+1; yellow if count >0 and < i+1; red if count ==0
        ws_d.conditional_formatting.add(cell, FormulaRule(formula=[f"Practice!$B$20>={i+1}"], fill=green_fill))
        ws_d.conditional_formatting.add(cell, FormulaRule(formula=[f"AND(Practice!$B$20>0,Practice!$B$20<{i+1})"], fill=yellow_fill))
        ws_d.conditional_formatting.add(cell, FormulaRule(formula=[f"Practice!$B$20=0"], fill=red_fill))

    # Charts: Expense Mix
    ws_d["F5"]="Expense"; ws_d["G5"]="Amount"
    exp_labels = ["Salaries","Rent","Office Supplies","Royalties","Advertising Royalties","Misc / Shortages"]
    exp_refs   = [sal, rent, sup, roy, adv, misc]
    r0=6
    for i,(lbl,refc) in enumerate(zip(exp_labels, exp_refs)):
        ws_d[f"F{r0+i}"]=lbl; ws_d[f"G{r0+i}"]=f"={refc}"
    bar = BarChart(); bar.title="Expense Mix"; bar.y_axis.title="Amount"
    vals = Reference(ws_d, min_col=7, min_row=r0, max_row=r0+len(exp_labels)-1)
    cats = Reference(ws_d, min_col=6, min_row=r0, max_row=r0+len(exp_labels)-1)
    bar.add_data(vals, titles_from_data=False); bar.set_categories(cats); bar.height=7; bar.width=12
    ws_d.add_chart(bar, "E6")

    # Net Margin Doughnut
    ws_d["J6"]="Metric"; ws_d["K6"]="Value"
    ws_d["J7"]="Net Margin %"; ws_d["K7"]=f"=MAX(0,MIN(100,{nim}))"
    ws_d["J8"]="Remainder";    ws_d["K8"]="=MAX(0,100-K7)"
    dough = DoughnutChart(); dough.title="Net Margin Gauge"; dough.height=8; dough.width=8
    data = Reference(ws_d, min_col=11, min_row=7, max_row=8)
    dough.add_data(data, titles_from_data=False)
    ws_d.add_chart(dough, "J9")

    # Navigation
    ws_d["A2"]="Go to Inputs";    ws_d["A2"].hyperlink="#'Inputs'!A1";    ws_d["A2"].style="Hyperlink"
    ws_d["B2"]="Go to Practice";  ws_d["B2"].hyperlink="#'Practice'!A1";  ws_d["B2"].style="Hyperlink"
    ws_d["C2"]="Go to ProTips";   ws_d["C2"].hyperlink="#'ProTips'!A1";   ws_d["C2"].style="Hyperlink"
    ws_d["D2"]="Go to Report";    ws_d["D2"].hyperlink="#'Report'!A1";    ws_d["D2"].style="Hyperlink"

    ws_d.sheet_view.showGridLines=False
    ws_d.freeze_panes="A3"

    # -------------------- Practice --------------------
    ws_pr = wb.create_sheet("Practice")
    ws_pr["A1"]="Practice Prompts"; ws_pr["A1"].font=brand_header_font(size=14)
    prompts = [
        ("Increase return count by 10% — note the change in Net Income.", ""),
        ("Raise ANF by $5 — what happens to Net Margin %?", ""),
        ("Cost per Return is Yellow — which two expenses would you reduce first?", ""),
        ("Reduce Advertising by 2% — what is the trade‑off to returns?", ""),
        ("Compare Good vs Best — which is realistic for 2026 and why?", ""),
    ]
    ws_pr["A3"]="Question"; ws_pr["B3"]="Your notes"; ws_pr["C3"]="Done?"
    ws_pr["A3"].font=ws_pr["B3"].font=ws_pr["C3"].font=brand_font(bold=True)
    r=4
    for q,_ in prompts:
        ws_pr[f"A{r}"]=q; ws_pr[f"B{r}"]=""  # response cell
        # Auto check: mark "✅" if B cell has any text
        ws_pr[f"C{r}"]=f'=IF(LEN(B{r})>0,"✅","")'
        r+=1

    # Completion count
    ws_pr["A20"]="Completed count"; ws_pr["B20"]=f"=COUNTIF(C4:C8,\"✅\")"

    # Full progress bar (5 segments) at row 2
    ws_pr["E2"]="Practice Progress"; ws_pr["E2"].font=brand_font(bold=True)
    for i in range(5):
        cell = f"{get_column_letter(6+i)}2"  # F2..J2
        ws_pr[cell] = ""
        ws_pr[cell].border = thin_border()
        ws_pr.conditional_formatting.add(cell, FormulaRule(formula=[f"$B$20>={i+1}"], fill=PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")))
        ws_pr.conditional_formatting.add(cell, FormulaRule(formula=[f"AND($B$20>0,$B$20<{i+1})"], fill=PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")))
        ws_pr.conditional_formatting.add(cell, FormulaRule(formula=[f"$B$20=0"], fill=PatternFill(start_color=RED, end_color=RED, fill_type="solid")))

    # Completion message
    ws_pr["A22"]="Status"
    ws_pr["B22"]='=IF($B$20=5,"All practice questions complete! ✅","")'

    for col in range(1, 11): ws_pr.column_dimensions[get_column_letter(col)].width=30
    ws_pr.sheet_view.showGridLines=False

    # -------------------- ProTips --------------------
    ws_tip = wb.create_sheet("ProTips")
    ws_tip["A1"]="Automated Pro Tips"; ws_tip["A1"].font=brand_header_font(size=14)
    ws_tip["A3"]="Tip"; ws_tip["A3"].font=brand_font(bold=True)
    tips = [
        ('=IF(Results!B13<=Inputs!$B$23,"Net Income is negative — review fixed costs and staffing levels.","")'),
        ('=IF(Results!B15<Inputs!$B$20,"Net Margin is below caution — consider small ANF increase or reduce salaries %.","")'),
        ('=IF(Inputs!B9>30,"Salaries exceed 30% of gross — consider part‑time staffing or schedule optimization.","")'),
        ('=IF(Inputs!B10>20,"Rent above 20% of gross — evaluate space optimization or renegotiation.","")'),
        ('=IF(Results!B14>Inputs!$B$18,"Cost/Return is high — check discounts %, supplies, and local advertising ROI.","")')
    ]
    r=4
    for t in tips: ws_tip[f"A{r}"]=t; r+=1
    ws_tip.column_dimensions["A"].width=120
    ws_tip.sheet_view.showGridLines=False

    # -------------------- Report --------------------
    ws_rep = wb.create_sheet("Report")
    ws_rep["A1"]="Liberty Tax — P&L Budget & Forecast Summary"; ws_rep["A1"].font=brand_header_font(size=14)
    ws_rep["A3"]="Region"; ws_rep["B3"]="=Welcome!B3"
    ws_rep["A4"]="Scenario"; ws_rep["B4"]="=Inputs!B1"
    ws_rep["A6"]="Net Income"; ws_rep["B6"]="=Results!B13"
    ws_rep["A7"]="Net Margin %"; ws_rep["B7"]="=Results!B15"
    ws_rep["A8"]="Cost per Return"; ws_rep["B8"]="=Results!B14"
    ws_rep["A10"]="Practice Progress"; ws_rep["B10"]="=Practice!B20"
    ws_rep["A11"]="Status"; ws_rep["B11"]='=IF(Practice!B20=5,"All practice complete ✅","In progress")'
    for c in range(1,4): ws_rep.column_dimensions[get_column_letter(c)].width=36
    ws_rep.sheet_view.showGridLines=False

    # Quick nav on every sheet footer-like spot
    for ws in [ws_w, ws_in, ws_p, ws_r, ws_d, ws_pr, ws_tip, ws_rep]:
        ws["L1"]="Home"; ws["L1"].hyperlink="#'Welcome'!A1"; ws["L1"].style="Hyperlink"

    return wb

def main():
    wb = build_wb()
    os.makedirs("dist", exist_ok=True)
    out = os.path.join("dist","LT_PnL_Tool_v0.5.xlsx")
    wb.save(out)
    print(out)

if __name__ == "__main__":
    main()
